import re
import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles import Border, Side
from io import BytesIO
from datetime import datetime, date
from zoneinfo import ZoneInfo
from collections import OrderedDict
from copy import copy

CENTER = Alignment(horizontal="center", vertical="center")

BORDER = Border(
    right = Side(style="thin"),
    top = Side(style="thin"),
    bottom = Side(style="thin"),
)

ERROR_HIGHLIGHT = PatternFill(
    fill_type="solid",
    start_color="FF4A0B",
    end_color="FF4A0B",
)

def main():
    st.title("Sales & Stock Reconciliation Report Generator")
    
    st.header("📘 Introduction")
    with st.expander("Click to expand"):
        st.write("""
                Upload the requriment files to the corresponding box and 
                provide the information to generate the summary excel sheet.

                Please ensure all the uploaded file is in excel format.
                
                Try download the excel file again if first download is fail. 
                Do not need to refresh the page.
                """)

    st.divider()
    ExcelUploadSection()
    choice = GetUserCompanyChoice()

    company_pages = {
        "ร้านย่อย": ThaiName,
        "GBH": GBH,
        "DH": DH,
        "HP": HP,
    }

    if choice in company_pages:
        st.divider()
        company_pages[choice]()

# region --- Entrance function for different companies with specific programme logic ---

def ThaiName():
    express_file = st.session_state.get("excel_file_1")
    stock_file = st.session_state.get("excel_file_2")

    if express_file is not None and stock_file is not None:
        express_data, bill_numbers, total = GetExpressData(express_file)
        express_data = SummariseByBarcode(express_data)
        stock_data = GetStockData(stock_file, 0)

        excel_file = GenerateExcel()
        excel_file = UpdateUserInputTitle(excel_file)
        excel_file = GetDateTime(excel_file)
        excel_file = GetBranchNumberAndVersion(excel_file)
        excel_file = UpdateBillNumberAndTotalProfit(excel_file, bill_numbers, total)
        excel_file = WriteMainData(excel_file, express_data, stock_data)
        excel_file = AdjustExcelColWidthAndAddBorder(excel_file)
        
        DownloadFile(excel_file)

def GBH():
    express_file = st.session_state.get("excel_file_1")
    stock_file = st.session_state.get("excel_file_2")

    if express_file is not None and stock_file is not None:
        start_date, end_date = GetUserInputDates()
        excel_file, option = GetTemplate("GBH")
    
        express_data, bill_numbers, total = GetExpressData(express_file)
        express_data = SummariseByBarcode(express_data)
        stock_data = GetStockData(stock_file, 1, option)

        excel_file = WriteGBHFileInformation(excel_file, start_date, end_date, bill_numbers, total)
        excel_file = WriteGBHFileMainData(excel_file, express_data, stock_data)
        
        DownloadFile(excel_file)

def DH():
    express_file = st.session_state.get("excel_file_1")
    stock_file = st.session_state.get("excel_file_2")

    if express_file is not None and stock_file is not None:
        start_date, end_date = GetUserInputDates()
        excel_file, option = GetTemplate("DH")

        express_data, bill_numbers, total = GetExpressData(express_file)
        express_data = SummariseByBarcode(express_data)
        stock_data = GetStockData(stock_file, 4, option)

        excel_file = WriteDHFileInformation(excel_file, start_date, end_date, bill_numbers, total)
        excel_file = WriteDHFileMainData(excel_file, express_data, stock_data)
        
        DownloadFile(excel_file)

def HP():
    express_file = st.session_state.get("excel_file_1")
    stock_file = st.session_state.get("excel_file_2")

    if express_file is not None and stock_file is not None:
        start_date, end_date = GetUserInputDates()
        excel_file, option = GetTemplate("HP")

        express_data, bill_numbers, total = GetExpressData(express_file)
        express_data = SummariseByBarcode(express_data)
        stock_data = GetStockData(stock_file, 5, option)

        excel_file = WriteHPFileInformation(excel_file, start_date, end_date, bill_numbers, total)
        excel_file = WriteHPFileMainData(excel_file, express_data, stock_data)
        
        DownloadFile(excel_file)

# endregion

# region --- Excel generation helper functions for company in Thai ---

def GenerateExcel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Row 1
    ws.merge_cells("A1:G1")
    ws["A1"] = "My Big Title"
    ws["A1"].alignment = CENTER

    # Row 2
    ws["A2"] = "บิล:"
    ws["A2"].font = Font(size=13, bold=True, color="9933FF")
    ws["A2"].fill = PatternFill(
        fill_type="solid",
        start_color="E2EFDA",
        end_color="E2EFDA",
    )

    ws.merge_cells("B2:D2")
    ws["B2"] = "Row 2: A-D merged"

    ws["E2"] = "Row 2: E"
    ws["E2"].alignment = CENTER

    ws.merge_cells("F2:G2")
    ws["F2"] = "Row 2: F-G merged"

    # Row 3
    ws.merge_cells("A3:E3")
    ws["A3"] = "Row 3: A-E merged"

    ws.merge_cells("F3:G3")
    ws["F3"] = "Row 3: F-G merged"
    ws["F3"].alignment = CENTER

    # Row 4
    ws.merge_cells("A4:D4")
    ws["A4"] = "Row 4: A-D merged"

    ws.merge_cells("E4:G4")
    ws["E4"] = "Row 4: E-G merged"

    ws["A5"] = "NO."
    ws["B5"] = "บาร์โค้ด"
    ws["C5"] = "ชื่อสินค้า"
    ws["D5"] = "จำนวน"
    ws["E5"] = "STOCK"
    ws["F5"] = "แพ็ค"
    ws["G5"] = "จัดสินค้า"

    for row in ws["A5:G5"]:
        for cell in row:
            cell.alignment = CENTER
            cell.font = Font(size=12, bold=True)

    ws.freeze_panes = "A6"

    return wb

def UpdateUserInputTitle(wb):
    """
    Shows a text box. Whatever the user types is stored automatically
    (no save button) and returned exactly as entered.
    """

    ws = wb.active

    # Always store the latest raw text
    ws["A1"].value = GetUserInputTitle()
    ws["A1"].font = Font(size=32, color="6600CC")
    ws["A1"].fill = PatternFill(
        fill_type="solid",
        start_color="FFAAFF",
        end_color="FFAAFF",
    )

    return wb

def GetDateTime(wb):
    """
    Show date & time inputs for the Excel file.
    - Defaults to current date & time on first run
    - User can edit any part they want
    - Values are stored live in st.session_state["date"] and ["time"]
    - No need to return anything; you can read from session_state later.
    """
    st.subheader("Date & Time")
    ws = wb.active

    now = datetime.now(ZoneInfo("Asia/Bangkok"))

    # Set defaults only once (so user edits are not overwritten on rerun)
    if "date" not in st.session_state:
        st.session_state["date"] = now.strftime("%Y-%m-%d")   # e.g. 2026-01-07
    if "time" not in st.session_state:
        st.session_state["time"] = now.strftime("%H:%M")   # e.g. 14:32:05

    date_val = st.date_input("Date", value=now.date(), key="date_input")
    time_val = st.time_input("Time", value=now.time(), key="time_input", step=300)

    st.session_state["date"] = date_val.strftime("%d/%m/") + str(date_val.year + 543)
    st.session_state["time"] = time_val.strftime("%H:%M")

    ws["E2"] = st.session_state["time"]
    ws["E2"].font = Font(size=14, color="000000")
    ws["E2"].fill = PatternFill(
        fill_type="solid",
        start_color="FFC000",
        end_color="FFC000",
    )

    ws["F2"] = "วันที่   " + st.session_state["date"]
    ws["F2"].font = Font(size=16, color="FF0000")
    ws["F2"].fill = PatternFill(
        fill_type="solid",
        start_color="FCE4D6",
        end_color="FCE4D6",
    )

    return wb

def GetBranchNumberAndVersion(wb):
    """
    Get the branch number and the version of the file 
    from user input and put in the excel
    """
    st.subheader("Branch Number & Version")
    ws = wb.active

    st.text_input(
        "Enter the branch number:",
        key = "branch_number",
        placeholder="Enter the branch number here",
    )

    st.text_input(
        "Enter the version:",
        key = "version",
        placeholder="Enter the version number here",
    )

    ws["A3"] = "เขต:  " + st.session_state["branch_number"]
    ws["A3"].font = Font(size=18, color="CC00FF")
    ws["A3"].fill = PatternFill(
        fill_type="solid",
        start_color="FFCCFF",
        end_color="FFCCFF",
    )

    ws["F3"] = st.session_state["version"]
    ws["F3"].font = Font(size=21, bold=True, color="0000FF")
    ws["F3"].fill = PatternFill(
        fill_type="solid",
        start_color="97DCFF",
        end_color="97DCFF",
    )
    #97DCFF

    return wb

def UpdateBillNumberAndTotalProfit(wb, bill_numbers, total):
    ws = wb.active

    bill_number_range = FindBillNumberRange(bill_numbers)
    ws["B2"] = bill_number_range
    ws["B2"].font = Font(size=20, color="0000FF")
    ws["B2"].fill = PatternFill(
        fill_type="solid",
        start_color="E2EFDA",
        end_color="E2EFDA",
    )

    ws["A4"] = "รวม                                         " + total + "   บาท"
    ws["A4"].font = Font(size=16, color="0066FF")
    ws["A4"].fill = PatternFill(
        fill_type="solid",
        start_color="CCCCFF",
        end_color="CCCCFF",
    )

    ws["E4"] = "จำนวนบิล         " + str(len(bill_numbers)) + "    บิล"
    ws["E4"].font = Font(size=14, color="FF0066")
    ws["E4"].fill = PatternFill(
        fill_type="solid",
        start_color="E2EFDA",
        end_color="E2EFDA",
    )

    return wb

def WriteMainData(wb, express_data, stock_data):
    ws = wb.active

    stock_lookup = {
        SafeInt(row[0]): row[1:]
        for row in stock_data
        if SafeInt(row[0]) is not None
    }

    for idx, item in enumerate(express_data, start=1):
        excel_row = idx + 5

        cell = ws[f"A{excel_row}"]
        cell.value = idx
        cell.alignment = CENTER
        cell.font = Font(size=12)

        ws[f"D{excel_row}"].value = item["sum_qty"]
        ws[f"D{excel_row}"].alignment = CENTER
        ws[f"D{excel_row}"].font = Font(size=12)

        if "_" in item["barcode"]:
            before, _, after = item["barcode"].partition("_")
            ws[f"B{excel_row}"].value = before
            ws[f"B{excel_row}"].alignment = CENTER
            ws[f"B{excel_row}"].font = Font(size=12)
            ws[f"B{excel_row}"].fill = ERROR_HIGHLIGHT

            ws[f"C{excel_row}"] = after
            ws[f"C{excel_row}"].alignment = CENTER
            ws[f"C{excel_row}"].font = Font(size=12)
            continue

        ws[f"B{excel_row}"].value = item["barcode"]
        ws[f"B{excel_row}"].alignment = CENTER
        ws[f"B{excel_row}"].font = Font(size=12)

        barcode = SafeInt(item.get("barcode"))
        stock_item = stock_lookup.pop(barcode, None)

        if stock_item is not None:
            ws[f"C{excel_row}"].value = stock_item[0]
            ws[f"C{excel_row}"].alignment = CENTER
            ws[f"C{excel_row}"].font = Font(size=12)

            ws[f"E{excel_row}"].value = stock_item[1]
            ws[f"E{excel_row}"].alignment = CENTER
            ws[f"E{excel_row}"].font = Font(size=12)
        else:
            ws[f"C{excel_row}"].value = (
                "Cannot find the barcode.\nUpdate the main sheet of the stock file."
            )
            ws[f"C{excel_row}"].alignment = CENTER
            ws[f"C{excel_row}"].font = Font(size=12)
            ws[f"C{excel_row}"].fill = ERROR_HIGHLIGHT

    return wb

def AdjustExcelColWidthAndAddBorder(wb):
    ws = wb.active

    # Choose the row and column want to autosize
    min_col = 1
    max_col = ws.max_column
    min_row = 5
    max_row = ws.max_row

    for col in range(min_col, max_col + 1):
        col_letter = get_column_letter(col)
        max_len = 0

        for row in range(min_row, max_row + 1):
            value = ws.cell(row=row, column=col).value
            ws.cell(row=row, column=col).border = BORDER
            if value is None:
                continue

            # Convert to string for measuring
            text = str(value)

            # Optional: ignore very long multi-line values
            text = text.split("\n")[0]

            if len(text) > max_len:
                max_len = len(text)

        # Add padding, and cap width so it doesn't get crazy wide
        padding = 6 if col != 1 else 3
        ws.column_dimensions[col_letter].width = min(max_len + padding, 60)

    return wb

#endregion

# region --- Excel generation helper functions for other companies ---

def GetTemplate(file_choice):
    templates = {
        "GBH": {"path": "template file/GBH.xlsx", "sheets": ["AS", "GL"]},
        "DH": {"path": "template file/DH.xlsx", "sheets": ["GL", "MR"]},
        "HP": {"path": "template file/HP.xlsx", "sheets": ["HP"]}
    }

    if file_choice not in templates:
        return None

    template = templates[file_choice]
    wb = load_workbook(template["path"])

    if len(template["sheets"]) == 1:
        sheet_choice = template["sheets"][0]
    else:
        st.subheader("Select sheet")
        sheet_choice = st.radio(
            "Select sheet",
            template["sheets"],
            index = 0,
            horizontal = True,
            key = f"{file_choice}_sheet_choice",
            label_visibility = "collapsed"
        )

    if not sheet_choice:
            return None
        
    for ws in wb.worksheets[:]:
            if ws.title != sheet_choice:
                wb.remove(ws)

    wb.active = 0
    return wb, sheet_choice

def WriteGBHFileInformation(wb, start_date, end_date, bill_number, total):
    ws = wb.active

    ws["F2"].value = start_date
    ws["G2"].value = end_date

    bill_number_range = FindBillNumberRange(bill_number)
    information = [bill_number_range, str(len(bill_number)), str(total)]
    for replacement in information:
        ws["A3"].value = ws["A3"].value.replace("?", replacement, 1)

    return wb

def WriteGBHFileMainData(wb, express_data, stock_data):
    return WriteExcelMainData(wb, express_data, stock_data)

def WriteDHFileInformation(wb, start_date, end_date, bill_number, total):
    ws = wb.active

    ws["I1"].value = ws["I1"].value.replace("?", start_date.replace(".", "/"))
    ws["A2"].value = start_date
    ws["E2"].value = end_date

    ws["A3"].value = ws["A3"].value.replace("?", str(total))
    ws["E3"].value = ws["E3"].value.replace("?", str(len(bill_number)))

    bill_number_range = FindBillNumberRange(bill_number)
    ws["A4"].value = ws["A4"].value.replace("?", bill_number_range)

    return wb

def WriteDHFileMainData(wb, express_data, stock_data):
    return WriteExcelMainData(wb, express_data, stock_data)

def WriteHPFileInformation(wb, start_date, end_date, bill_number, total):
    ws = wb.active

    information = [start_date, end_date, str(total)]
    for replacement in information:
        ws["A1"].value = ws["A1"].value.replace("?", replacement, 1)

    bill_number_range = FindBillNumberRange(bill_number)
    ws["A2"].value = (
        ws["A2"].value
        .replace("?", bill_number_range, 1)
        .replace("?", str(len(bill_number)), 1)
    )

    return wb

def WriteHPFileMainData(wb, express_data, stock_data):
    return WriteExcelMainData(wb, express_data, stock_data)

def WriteExcelMainData(wb, express_data, stock_data):
    ws = wb.active

    header_end_row = GetLastRealRow(ws)
    stock_col =  8 if any(
        isinstance(ws.cell(row=header_end_row, column=col).value, str)
        and "stock" in ws.cell(row=header_end_row, column=col).value.lower()
        for col in range(1, ws.max_column + 1)
    ) else 7
    
    stock_lookup = {}
    for s in stock_data:
        barcode = SafeInt(s[0])
        if not barcode:
            continue

        # Always: (detail, info, stock)
        stock_lookup[barcode] = (
            s[1],
            s[2] if len(s) >= 4 else None,
            s[-1],
        )

    column_styles = CaptureColumnStyles(ws, header_end_row+1)
    sum = 0.0

    for idx, item in enumerate(express_data, start = 1):
        write_row = header_end_row + idx

        index_cell = ws.cell(row=write_row, column = 1)
        barcode_cell = ws.cell(row=write_row, column=2)
        detail_cell = ws.cell(row=write_row, column=3)
        addition_info_cell = ws.cell(row=write_row, column=4)
        amount_cell = ws.cell(row=write_row, column=5)
        stock_cell = ws.cell(row=write_row, column=stock_col)

        for col, style in column_styles.items():
            ApplyColumnStyleToCell(ws.cell(write_row, col), style)

        index_cell.value = idx
        amount_cell.value = item["sum_qty"]
        sum += item["sum_qty"]

        barcode = item.get("barcode", "")
        if "_" in barcode:
            barcode_cell.value, detail_cell.value = barcode.split("_", 1)
            barcode_cell.fill = ERROR_HIGHLIGHT
            continue

        barcode_cell.value = barcode
        stock_item = stock_lookup.pop(SafeInt(barcode), None)

        if stock_item is not None:
            detail, info, stock = stock_item
            detail_cell.value = detail
            stock_cell.value = stock

            if info is not None:
                addition_info_cell.value = info
        else:
            detail_cell.value = "Cannot find the barcode.\nUpdate the main sheet."
            detail_cell.fill = ERROR_HIGHLIGHT

    cell = ws[f"E{GetLastRealRow(ws)+1}"]
    cell.value = sum
    cell.font = copy(ws[f"E{GetLastRealRow(ws)}"].font) + Font(bold=True)
    cell.fill = PatternFill(
        fill_type="solid",
        start_color="FFFF00",
        end_color="FFFF00",
    )

    AutoResizeColumn(ws, 3, end_row=GetLastRealRow(ws)-1, padding=0, max_width=90)

    return wb

def CaptureColumnStyles(ws, style_row):
    styles = {}
    max_column = GetLastRealCol(ws)
    for col in range(1, max_column + 1):
        cell = ws.cell(row=style_row, column=col)
        styles[col] = {
            "font": copy(cell.font),
            "border": copy(cell.border),
            "fill": copy(cell.fill),
            "number_format": cell.number_format,
            "alignment": copy(cell.alignment),
            "protection": copy(cell.protection),
        }
    return styles

def ApplyColumnStyleToCell(cell, style):
    cell.font = copy(style["font"])
    cell.border = copy(style["border"])
    cell.fill = copy(style["fill"])
    cell.number_format = style["number_format"]
    cell.alignment = copy(style["alignment"])
    cell.protection = copy(style["protection"])
    cell.border = BORDER

def AutoResizeColumn(ws, col, start_row=1, end_row=None, 
                     padding=2, min_width=8, max_width=50):
    if end_row is None:
        end_row = ws.max_row

    max_width_est = 0
    merged_ranges = ws.merged_cells.ranges
    col_letter = get_column_letter(col)

    for row in range(start_row, end_row + 1):
        coord = f"{col_letter}{row}"

        skip = False
        for merged in merged_ranges:
            if coord in merged and not (row == merged.min_row and col == merged.min_col):
                skip = True
                break
        if skip:
            continue

        cell = ws.cell(row=row, column=col)
        if not cell.value:
            continue

        text = str(cell.value)
        font_size = cell.font.sz or 11
        scale = font_size / 11

        est = int(len(text) * scale) + padding
        max_width_est = max(max_width_est, est)

    final_width = max(min_width, min(max_width_est, max_width))
    ws.column_dimensions[col_letter].width = final_width

#endregion

# region --- Data obtain & analysis helper functions ---

def GetExpressData(uploaded_file):
    """
    Given an uploaded Excel file, find the SECOND 'horizontal line' row
    (a row where any cell contains only '-' characters like '--------')
    and return all rows of data below that line.

    Returns:
        list[list]: list of rows; each row is a list of cell values.
    """
    # Make sure we're at the start of the file
    try:
        uploaded_file.seek(0)
    except Exception:
        # Some file-like objects may not have seek; ignore if so
        pass

    wb = load_workbook(uploaded_file, data_only=True)
    ws = wb.active  # or wb["SheetName"] if you want a specific sheet

    max_row = ws.max_row
    max_col = ws.max_column

    separator_rows = []

    for r in range(1, max_row + 1):
        row_has_dash_line = False

        for c in range(1, max_col + 1):
            cell_value = ws.cell(row=r, column=c).value

            if isinstance(cell_value, str):
                stripped = cell_value.strip()
                # check "purely horizontal line", e.g. "-----" or " -------- "
                if stripped and set(stripped) == {"-"}:
                    row_has_dash_line = True
                    break  # we already know this row is a separator

        if row_has_dash_line:
            separator_rows.append(r)

    # Need at least 2 separator rows
    if len(separator_rows) < 2:
        raise ValueError("Invalid Express File Format: missing seperators (second horizontal line)")

    second_sep_row = separator_rows[1]
    start_row = second_sep_row + 1

    data = []
    for r in range(start_row, max_row + 1):
        row_values = [
            ws.cell(row=r, column=c).value
            for c in range(1, max_col + 1)
        ]

        # Optionally skip completely empty rows or integer row
        if all(v in (None, "") for v in row_values) \
            or all(type(v) in (int, float) for v in row_values):
                continue

        split_row = row_values[0].split()
        data.append(split_row)

    return TreatExpressData(data)

def TreatExpressData(data):
    bill_number = ""
    bill_number_collection = []
    index = 0

    for row in data:
        if len(row) < 3:
            continue

        third = str(row[2]).strip()

        # already good
        if third.isdigit():
            continue

        # try to repair "1234.SOMETHING" => ["1234", "SOMETHING"]
        if "." in third:
            left, _, right = third.partition(".")

            # only split if left is digits AND right looks like a real next-field (not just decimals)
            if left.isdigit() and right and not right.isdigit():
                row[2] = left
                row.insert(3, right)  # shift the rest to the right
                continue

        if not third.isdigit():
            row[2] = "0000000000000_"+ third
            row.insert(3, third)

    while (index < len(data) and data[index][0] != "รวมทั้งสิ้น"):
        if len(data[index]) < 5 or has_thai(data[index][0]):
            del data[index]
            continue
        
        checking_bill = data[index][0]
        checking_bill = re.sub(r'[^A-Za-z0-9]', '', checking_bill)

        if (checking_bill != bill_number):
            bill_number = checking_bill
            bill_number_collection.append(bill_number)
            del data[index]
            index -= 1
    
        index += 1

    if (data[index][0] == "รวมทั้งสิ้น"):
            total = (data[index][-3])
            data = data[0 : index]
            return data, bill_number_collection, total  

    raise ValueError("Cannot find รวมทั้งสิ้น, check the input file.")

def FindBillNumberRange(bill_number):
    """
    Convert a list of alphanumeric bill numbers into compact string ranges.
    Continuous numeric parts are compressed with "-", gaps separated by "/".
    Original prefixes are preserved.
    """
    if not bill_number:
        return ""

    # Step 1: Extract numeric part and keep original ID mapping
    id_map = []
    for i in bill_number:
        num = int(re.sub(r"\D", "", i))  # numeric part
        id_map.append((num, i))          # tuple (number, original_id)

    # Step 2: Sort by numeric part
    id_map.sort(key=lambda x: x[0])

    # Step 3: Build ranges
    ranges = []
    start_num, start_id = id_map[0]
    end_num, end_id = start_num, start_id

    for num, original_id in id_map[1:]:
        if num == end_num + 1:
            # continuous
            end_num = num
            end_id = original_id
        else:
            # gap found, save previous range
            if start_num == end_num:
                ranges.append(start_id)
            else:
                ranges.append(f"{start_id}-{end_id}")
            start_num, start_id = num, original_id
            end_num, end_id = num, original_id

    # Add the last range
    if start_num == end_num:
        ranges.append(start_id)
    else:
        ranges.append(f"{start_id}-{end_id}")

    # Join ranges with /
    return " / ".join(ranges)

def ExtractPackQtyFromRow(row):
    """
    Find all cells in the row that contain '.แพ็ค' and
    sum the numbers before '.แพ็ค'.

    E.g. '55.แพ็ค' -> 55, '8.แพ็ค' -> 8.
    If nothing found or parse fails, returns 0.
    """
    qty = 0.0
    suffixes = (
        ".แพ็ค", ".ชิ้น", ".อัน", ".ชุด", ".แผ่น",
        ".กล่อง", ".ถุง", ".ม้วน", ".ลัง", ".แผง", ".คู่",
        ".เครื่อง", ".ขวด", ".กระป๋อง", ".เส้น", ".ตัว", ".ใบ",
        ".เมตร", ".ลูก", ".โหล", ".ดวง"
    )

    found = False
    for cell in row:
        if isinstance(cell, str) and any(suffix in cell for suffix in suffixes):
            # Find the first matching suffix
            for suffix in suffixes:
                if suffix in cell:
                    before = cell.split(suffix)[0].strip()
                    break

            before = before.replace(",", "")
            if not before:
                continue

            qty += float(before)
            found = True

    if not found:
        raise ValueError("Self Defined Error 10010: No valid suffix found in this row!")  # raise Python exception

    return qty

def SummariseByBarcode(data_rows):
    """
    Group rows by (barcode, item_code) = (row[2], row[3])
    and sum all 'X.แพ็ค' quantities for each group.

    Returns:
        list of dicts:
          {
            'barcode': ...,
            'item_code': ...,
            'sum_qty': ...,
          }
    """
    summaries = OrderedDict()  # to keep order of first appearance

    for row in data_rows:
        if len(row) < 4:
            continue  # need at least [bill, line, barcode, item_code]

        barcode = row[2]
        if barcode is None:
            continue

        if barcode not in summaries:
            summaries[barcode] = {
                "barcode": barcode,
                "sum_qty": 0.0,
            }

        qty = ExtractPackQtyFromRow(row)
        summaries[barcode]["sum_qty"] += qty

    return list(summaries.values())

def GetStockData(uploaded_file, sheet, option=None):
    """
    Given the uploaded stock Excel file, search the barcode that listed before through the file.
    (which is appear at the second column in the stock file)
    and return all rows of barcode with the stock number (stored in the sixth column)

    Returns:
        List: a list of the stock in the order of barcode searching order.
    """
    # Make sure we're at the start of the file
    try:
        uploaded_file.seek(0)
    except Exception:
        # Some file-like objects may not have seek; ignore if so
        pass

    if sheet != 0 and option == "MR":
        data_cols = [2, 3, 4, 5]
    elif sheet != 0 and option != "GL":
        data_cols = [2, 3, 4, 6]
    else:
        data_cols = [2, 3, 6]
        sheet = {1: 2, 4: 3}.get(sheet, sheet)

    wb = load_workbook(uploaded_file, data_only=True)
    ws = wb.worksheets[sheet]  # or wb["SheetName"] if you want a specific sheet

    max_row = ws.max_row
    data = []

    for r in range (2, max_row+1):
        row_value = [
            ws.cell(row=r, column=c).value
            for c in data_cols
        ]

        # Optionally skip completely empty rows
        if all(v in (None, "") for v in row_value):
            continue
        
        data.append(row_value)

    return data

#endregion

# region --- General user interface functions ---

def ExcelUploadSection():
    """
    Show an interface that lets the user upload two Excel files.
    The uploaded files are stored in:
        st.session_state["excel_file_1"]
        st.session_state["excel_file_2"]
    so you can use them later in your code.
    """

    st.subheader("Upload Excel Files")

    file1 = st.file_uploader(
        "Upload the Excel file from Express Accounting.",
        type=["xlsx", "xlsm", "xls"],
        key=f"excel_upload_1",
    )

    file2 = st.file_uploader(
        "Upload the product stock file",
        type=["xlsx", "xlsm", "xls"],
        key=f"excel_upload_2",
    )

    # Store them in session_state so other blocks can use them
    if file1 is not None:
        st.session_state["excel_file_1"] = file1

    if file2 is not None:
        st.session_state["excel_file_2"] = file2

def GetUserCompanyChoice():
    st.subheader("กรุณาเลือกลูกค้า")

    if "prev_choice" not in st.session_state:
        st.session_state.prev_choice = None

    options = ["ร้านย่อย", "GBH", "DH", "HP"]
    cols = st.columns(len(options))

    for col, option in zip(cols, options):
        is_selected = option == st.session_state.prev_choice
        label = f"✅ {option}" if is_selected else option

        if col.button(label, use_container_width=True):
            if option != st.session_state.prev_choice:
                keep_keys = {"excel_file_1", "excel_file_2", "prev_choice"}
                for key in list(st.session_state.keys()):
                    if key not in keep_keys:
                        del st.session_state[key]

                st.session_state.prev_choice = option
                st.session_state.toggle_button = False
                st.rerun()

    return st.session_state.prev_choice

def GetUserInputTitle():
    st.subheader("Title")
    title = st.text_input(
        "Enter title:",
        key = "user_title",
        placeholder="Enter the title here",
        label_visibility="collapsed"
    )

    return title

def GetUserInputDates():
    today = date.today()

    st.subheader("Choose Date range")
    date_range = st.date_input(
        "Select date range",
        value = (today, today),
        label_visibility="collapsed"
    )

    if isinstance(date_range, tuple):
        # If the user fortgot to choice the end date, set the today as end date
        if len(date_range) == 1:
            start_date = date_range[0]
            end_date = today
        else:
            start_date, end_date = date_range
    else:
        start_date = end_date = date_range

    # Change the returned date format
    start_date = f"{start_date.strftime('%d.%m')}.{start_date.year + 543}"
    end_date = f"{end_date.strftime('%d.%m')}.{end_date.year + 543}"
    return start_date, end_date

def DownloadFile(wb):
    st.divider()
    st.subheader("Download the Excel file")

    agree = st.toggle(
        "I confirm I am not a robot",
        key="toggle_button",
        value=False
    )

    output_excel_file = BytesIO()
    wb.save(output_excel_file)
    output_excel_file.seek(0)

    st.download_button(
        label="⬇️ Download Excel File",
        data=output_excel_file,
        file_name="output.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        disabled = not agree,
    )

#endregion

# region --- General helper function ---
   
def SafeInt(x):
    try:
        return int(x)
    except (ValueError, TypeError):
        return None

def GetLastRealRow(ws, col=1):
    row = ws.max_row
    while row > 0 and ws.cell(row=row, column=col).value is None:
        row -= 1
    return row

def GetLastRealCol(ws):
    col = ws.max_column
    while col > 0 and ws.cell(row=GetLastRealRow(ws), column=col).value is None:
        col -= 1
    return col

def has_thai(value) -> bool:
    if value is None:
        return False

    text = str(value)
    return any('\u0E00' <= ch <= '\u0E7F' for ch in text)

# endregion

main()
