import streamlit as st
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import column_index_from_string
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from PIL import Image
import io

st.title("Product Image Inserter")

template_file = st.file_uploader("Upload Excel Template (.xlsx)", type=["xlsx"])
product_images_file = st.file_uploader("Upload Product Images (.xlsx)", type=["xlsx"])

def insert_resized_image_center(ws, row, img_bytes):
    cell_height = ws.row_dimensions[row].height * 1.2
    cell_width = 300

    img = Image.open(io.BytesIO(img_bytes))
    w, h = img.size
    st.write(w, "and", h)
    ratio = min(cell_width/w, cell_height/h)
    img_resized = img.resize((int(w*ratio), int(h*ratio)))

    buffer = io.BytesIO()
    img_resized.save(buffer, format="PNG")
    buffer.seek(0)
    new_img = XLImage(buffer)

    ws.add_image(new_img, f"B{row}")

    col_px = int(ws.column_dimensions["B"].width * 7 + 5)
    row_px = ws.row_dimensions[row].height * 96 / 72
    off_x = (col_px - new_img.width) / 2 * 9525
    off_y = (row_px - new_img.height) / 2 * 9525
    if off_x < 0: off_x = 0
    if off_y < 0: off_y = 0

    _from = AnchorMarker(
        col = 1,
        row = row - 1,
        colOff = int(off_x),
        rowOff = int(off_y)
    )

    ext = XDRPositiveSize2D(
        cx = new_img.width * 9525,
        cy =  new_img.height * 9525
    )

    new_img.anchor = OneCellAnchor(_from = _from, ext = ext)


if template_file and product_images_file and st.button("Process"):
    template_wb = load_workbook(io.BytesIO(template_file.read()))
    template_ws = template_wb.active

    product_images_wb = load_workbook(io.BytesIO(product_images_file.read()))
    product_images_ws = product_images_wb.active

    template_ws.column_dimensions["B"].width = 30  # default width if not set

    for row in range(9, 14):
            product_number = str(template_ws[f"A{row}"].value)

            for product in range(3, product_images_ws.max_row+1):
                if product_number == str(product_images_ws[f"C{product}"].value):
                    for img in product_images_ws._images:
                        img_row = img.anchor._from.row + 1
                        img_col = img.anchor._from.col + 1
                        if img_row == product and img_col == 2:
                            image = img
                            break

                    img_bytes = image._data()
                    insert_resized_image_center(template_ws, row, img_bytes)

    output = io.BytesIO()
    template_wb.save(output)
    output.seek(0)

    st.success("Image copied to template successfully!")

    st.download_button(
        label="Download Updated Template",
        data=output,
        file_name="updated_template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
