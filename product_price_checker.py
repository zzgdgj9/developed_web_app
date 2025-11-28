# app.py
import streamlit as st
import pandas as pd
import re
from openpyxl import load_workbook
from io import BytesIO

st.title("Excel Matcher — STBALLL ↔ Update Price")

file_left = st.file_uploader("Upload STBALLL Excel/CSV file", type=["xlsx","xls","csv"])
file_right = st.file_uploader("Upload Update Price Excel file", type=["xlsx","xls"])

# =========================================================
# --- Improved Barcode Cleaning (ONLY this part changed) ---
# =========================================================

SPACE_CHARS = r"\u0020\u00A0\u1680\u2000-\u200A\u202F\u205F\u3000"

def clean_barcode(raw: str) -> str:
    if pd.isna(raw):
        return ""

    s = str(raw)

    # Normalize ALL weird spaces to normal space
    s = re.sub(f"[{SPACE_CHARS}]+", " ", s)

    s = s.strip()

    # Remove repeated "No", "NO", "no", etc. at the beginning
    s = re.sub(r"^(?:No)+", "", s, flags=re.IGNORECASE).strip()

    # Extract until space, slash, plus, Thai chars — KEEP brackets
    m = re.match(rf"^[^ /\+\u0E00-\u0E7F]+", s)
    return m.group(0) if m else s


# =========================================================
# --- Other helper functions (unchanged) ---
# =========================================================

def is_integer_token(tok: str) -> bool:
    if pd.isna(tok):
        return False
    return str(tok).strip().isdigit()

def numeric_value_for_compare(s: str):
    if pd.isna(s):
        return float('nan')
    t = re.sub(r'[^\d.\-]', '', str(s))
    try:
        return float(t)
    except:
        return float('nan')

def read_any_table(uploaded):
    fname = uploaded.name.lower()
    if fname.endswith((".xlsx",".xls")):
        return pd.read_excel(uploaded, header=None, dtype=str, engine="openpyxl")
    elif fname.endswith(".csv"):
        encodings = ["utf-8","utf-8-sig","cp1252","latin1","gbk","big5"]
        for enc in encodings:
            try:
                return pd.read_csv(uploaded, header=None, dtype=str, encoding=enc)
            except:
                continue
        raise ValueError("Could not read CSV file with common encodings")
    else:
        raise ValueError("Unsupported file type")

# =========================================================
# --- Main processing (unchanged except uses new cleaner) ---
# =========================================================

if st.button("Process files"):

    if file_left is None or file_right is None:
        st.error("Please upload both files.")
    else:
        st.info("Reading files...")
        left_df = read_any_table(file_left)
        right_df = pd.read_excel(file_right, header=0, dtype=str, engine="openpyxl")

        left_indices = []
        left_products = []
        left_prices = []

        for i, row in left_df.iterrows():
            row_str = str(row[0])
            cols = re.split(r'\s{2,}', row_str.strip())
            if len(cols) < 2:
                continue
            if not is_integer_token(cols[0]):
                continue

            left_indices.append(i)
            cleaned_prod = clean_barcode(cols[1])   # ← NEW CLEANER USED HERE
            left_products.append(cleaned_prod)

            if len(cols) >= 4 and numeric_value_for_compare(cols[3]) == numeric_value_for_compare(cols[3]):
                left_prices.append(cols[3])
            elif len(cols) >= 5:
                left_prices.append(cols[4])
            else:
                left_prices.append("")

        left_table = pd.DataFrame({
            "Index": left_indices,
            "Product": left_products,
            "Unit Price": left_prices
        })

        keep_unmatch_idx = []
        keep_outdated_idx = []

        for i, row in right_df.iterrows():
            search = clean_barcode(row.iloc[0])   # ← NEW CLEANER USED HERE
            found_idx = None

            for j, prod in left_table['Product'].items():
                if str(search).strip() == str(prod).strip():
                    found_idx = j
                    break

            if found_idx is None:
                keep_unmatch_idx.append(i)
            else:
                left_price_val = numeric_value_for_compare(left_table.loc[found_idx, 'Unit Price'])
                right_price_val = numeric_value_for_compare(row.iloc[3]) if len(row) > 3 else float('nan')
                if round(left_price_val,2) != round(right_price_val,2):
                    keep_outdated_idx.append(i)

        wb = load_workbook(file_right)
        original_sheet = wb.active

        sheet_unmatch = wb.copy_worksheet(original_sheet)
        sheet_unmatch.title = "Not Found Product"

        sheet_outdated = wb.copy_worksheet(original_sheet)
        sheet_outdated.title = "Outdated Unit Price"

        wb.remove(original_sheet)

        def hide_rows(sheet, keep_indices):
            total_rows = sheet.max_row
            keep_excel_rows = [i+2 for i in keep_indices]
            for r in range(2, total_rows+1):
                if r not in keep_excel_rows:
                    sheet.row_dimensions[r].hidden = True

        hide_rows(sheet_unmatch, keep_unmatch_idx)
        hide_rows(sheet_outdated, keep_outdated_idx)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        st.success("Processing complete. Download result:")
        st.download_button(
            label="Download result.xlsx",
            data=output,
            file_name="result.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )